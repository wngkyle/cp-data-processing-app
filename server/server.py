from flask import Flask, jsonify, request
from flask_cors import CORS
from openpyxl import Workbook
from main import reading_and_processing
from data import fileDetail, progressTracker, Isc_20mA_data, Turn_off_80mA_data, Turn_off_80mA_HL_data, Rf_data, Rr_data
from modified_step import Isc_20_mA_step_size, Turn_off_80mA_step_size, Turn_off_80mA_HL_step_size, Rf_step_size, Rr_step_size
import os, shutil
import matplotlib.pyplot as plt
import matplotlib

####################################################################################################################################
#                                                         SERVER                                                                   #
####################################################################################################################################

app = Flask(__name__)
CORS(
    app,
    origins=["http://localhost:3000"],
)

# @app.before_request
# def initialize():
#     plt.figure(figsize=(10, 6))

@app.route('/get-current-working-directory')
def getCurrWork():
    cwd = os.getcwd()
    return cwd

@app.route('/get-list-of-folders')
def getListOfDir():
    cwd = getCurrWork()
    listOfDir = os.scandir(cwd)
    listOfDir = check_files(listOfDir)
    return jsonify(listOfDir)

@app.route('/directory-backward')
def directoryBackward():
    os.chdir('../')
    cwd = os.getcwd()
    return cwd

@app.route('/directory-forward', methods=['POST'])
def directoryForward():
    data = request.json
    selectedFolder = data.get('selectedFolder')
    path = f'./{selectedFolder}'
    os.chdir(path)
    cwd = os.getcwd()
    return cwd

@app.route('/set-folder-and-create', methods=['POST'])
def setFolderAndCreate():
    data = request.json
    selectedFolder = data.get('finalFolder')
    cwd = os.getcwd()
    selectedFolderPath = f'{cwd}/{selectedFolder}'
    fileDetail['selected_folder'] = selectedFolderPath
    finalFolderPath = f'{fileDetail["selected_folder"]}_Processed'
    fileDetail['fdpath'] = finalFolderPath
    os.mkdir(fileDetail['fdpath'])
    return finalFolderPath

@app.route('/set-selected-columns', methods=['POST'])
def setSelectedColumns():
    data = request.json
    selectedColumns = data.get('selectedColumns')
    for col in list(selectedColumns.keys()):
        fileDetail['selected_column'][col] = 1
    setStepSize(selectedColumns)
    return selectedColumns

def setStepSize(selectedColumns):
    temp = list(selectedColumns.keys())
    if ('Isc_20mA' in temp):
        Isc_20mA_data['step_size_Isc_20mA'] = float(selectedColumns['Isc_20mA'])
    if ('Turn_off_80mA_' in temp):
        Turn_off_80mA_data['step_size_Turn_off_80mA'] = float(selectedColumns['Turn_off_80mA_'])
    if ('Turn_off_80mA_HL' in temp):
        Turn_off_80mA_HL_data['step_size_Turn_off_80mA_HL'] = float(selectedColumns['Turn_off_80mA_HL'])
    if ('Rf' in temp):
        Rf_data['step_size_Rf'] = float(selectedColumns['Rf'])
    if ('Rr' in temp):
        Rr_data['step_size_rr'] = float(selectedColumns['Rr'])


@app.route('/set-fast-track', methods=['POST'])
def setFastTrack():
    data = request.json
    fastTrackk = data.get('fastTrack')
    if fastTrackk == 'true':
        fileDetail['fast_track'] = True
    else:
        fileDetail['fast_track'] = False
    return fastTrackk

# TASKSSSS:
# Check if step size number needs to be converted into integer after sending it from frontend
# Check default, if users enter nothing, what is the default value for step size
    # might have to modify frontend as well, maybe show the default in the input tag as well
@app.route('/exec')
def exec():
    stage = []
    stage.append('**** Enter exec() in server ****')
    try :
        stage.append(' - File Creation Start')
        if (fileDetail['selected_column']['Isc_20mA'] == 1):
            Isc_20_mA_step_size()
            # Create a new file and new folder for storing charts
            fileDetail['isc_fdpath'] = f'{fileDetail["fdpath"]}/Isc_20mA'
            os.mkdir(fileDetail['isc_fdpath'])
            fileDetail['isc_fdpath_wmap'] = f'{fileDetail["fdpath"]}/Isc_20mA_wmap'
            os.mkdir(fileDetail['isc_fdpath_wmap'])
            wb = Workbook()
            wb.save(f'{fileDetail["fdpath"]}/Isc_20mA.xlsx')
            wb_wmap = Workbook()
            wb_wmap.save(f'{fileDetail["fdpath"]}/Isc_20mA_wmap.xlsx')
            stage.append(' - Isc_20mA File Creation Complete')

        if (fileDetail['selected_column']['Turn_off_80mA_'] == 1):
            Turn_off_80mA_step_size()
            # Create a new file and new folder for storing charts
            fileDetail['turn_off_fdpath'] = f'{fileDetail["fdpath"]}/Turn_off_80mA_'
            os.mkdir(fileDetail['turn_off_fdpath'])
            fileDetail['turn_off_fdpath_wmap'] = f'{fileDetail["fdpath"]}/Turn_off_80mA__wmap'
            os.mkdir(fileDetail['turn_off_fdpath_wmap'])
            wb = Workbook()
            wb.save(f'{fileDetail["fdpath"]}/Turn_off_80mA_.xlsx')
            wb_wmap = Workbook()
            wb_wmap.save(f'{fileDetail["fdpath"]}/Turn_off_80mA__wmap.xlsx')
            stage.append(' - Turn_off_80mA File Creation Complete')

        if (fileDetail['selected_column']['Turn_off_80mA_HL'] == 1):
            Turn_off_80mA_HL_step_size()
            # Create a new file and new folder for storing charts
            fileDetail['turn_off_HL_fdpath'] = f'{fileDetail["fdpath"]}/Turn_off_80mA_HL'
            os.mkdir(fileDetail['turn_off_HL_fdpath'])
            fileDetail['turn_off_HL_fdpath_wmap'] = f'{fileDetail["fdpath"]}/Turn_off_80mA_HL_wmap'
            os.mkdir(fileDetail['turn_off_HL_fdpath_wmap'])
            wb = Workbook()
            wb.save(f'{fileDetail["fdpath"]}/Turn_off_80mA_HL.xlsx')
            wb_wmap = Workbook()
            wb_wmap.save(f'{fileDetail["fdpath"]}/Turn_off_80mA_HL_wmap.xlsx')
            stage.append(' - Turn_off_80mA_HL File Creation Complete')

        if (fileDetail['selected_column']['Rf'] == 1):
            Rf_step_size()
            # Create a new file and new folder for storing charts
            fileDetail['rf_fdpath'] = f'{fileDetail["fdpath"]}/Rf'
            os.mkdir(fileDetail['rf_fdpath'])
            fileDetail['rf_fdpath_wmap'] = f'{fileDetail["fdpath"]}/Rf_wmap'
            os.mkdir(fileDetail['rf_fdpath_wmap'])
            wb = Workbook()
            wb.save(f'{fileDetail["fdpath"]}/Rf.xlsx')
            wb_wmap = Workbook()
            wb_wmap.save(f'{fileDetail["fdpath"]}/Rf_wmap.xlsx')
            stage.append(' - Rf File Creation Complete')

        if (fileDetail['selected_column']['Rr'] == 1):
            Rr_step_size()  
            # Create a new file and new folder for storing charts
            fileDetail['rr_fdpath'] = f'{fileDetail["fdpath"]}/Rr'
            os.mkdir(fileDetail['rr_fdpath'])
            fileDetail['rr_fdpath_wmap'] = f'{fileDetail["fdpath"]}/Rr_wmap'
            os.mkdir(fileDetail['rr_fdpath_wmap'])
            wb = Workbook()
            wb.save(f'{fileDetail["fdpath"]}/Rr.xlsx')
            wb_wmap = Workbook()
            wb_wmap.save(f'{fileDetail["fdpath"]}/Rr_wmap.xlsx')
            stage.append(' - Rr File Creation Complete')
    except:
        stage.append('Error in file creation: Error locate in first half of exec() in server.py')
    
    matplotlib.use('agg')
    try:
        stage.append(' - Calling reading_and_processing()')
        stage += reading_and_processing()
    except:
        stage.append('Error when preparing to enter reading_and_processing(): Error occur because of calling reading_and_processing() or in the function itself in server.py or main.py')
    stage.append('SUCCESSFUL')
    return stage
        
@app.route('/get-progress')
def getProgress():
    count = progressTracker['count']
    progressStatements = progressTracker['progressStatements']
    result = [count, progressStatements]
    return result

@app.route('/remove-processed-folder-content')
def removeProcessedFolderContent():
    finalFolderPath = fileDetail['fdpath']
    shutil.rmtree(finalFolderPath)
    os.mkdir(finalFolderPath)
    return 'Folder Content Deleted'

@app.route('/remove-folder')
def removeFolder():
    finalFolderPath = fileDetail['fdpath']
    os.rmdir(finalFolderPath)
    return 'Folder Deleted'
    
@app.route('/set-test', methods=['POST'])
def testRoute():
    data = request.json.get('data')
    return data

@app.route('/get-test')
def getTest():
    result = 'Test Successful'
    return result

@app.route('/check-file-detail')
def checkFileDetail():
    result = []
    result.append(fileDetail['selected_folder'])
    result.append(fileDetail['fdpath'])
    result.append(fileDetail['selected_column'])
    result.append(fileDetail['fast_track'])
    return result

@app.route('/test-plt')
def testPlt():
    matplotlib.use('agg')
    path = '/Users/wng/Desktop'
    x_values = [1, 2, 3, 4, 5]
    y_values = [2, 4, 6, 8, 10]

    # Create the plot
    plt.plot(x_values, y_values)

    # Add title and labels
    plt.title('Simple Line Chart')
    plt.xlabel('X-axis Label')
    plt.ylabel('Y-axis Label')

    # Display the plot
    plt.savefig(path + '/test.png')

    return 'sucessful'

def check_files(list):
    result = []
    for file in list:
        if file.is_dir() == False or file.name == '.git' or file.name == '__pycache__' or file.name == '$RECYCLE.BIN' or file.name == '.ipynb_checkpoints' or file.name == '.jpeg File':
            pass
        else:
            result.append(file.name)
    return result

if __name__ == '__main__':
    app.run(debug=True)