print('Start executing server code')

from flask import Flask, jsonify, request
from flask_cors import CORS
from openpyxl import Workbook
from main import reading_and_processing
from data import fileDetail, progressTracker, Isc_20mA_data, Turn_off_80mA_data, Turn_off_80mA_HL_data, Rf_data, Rr_data, processed
from modified_step import Isc_20_mA_step_size, Turn_off_80mA_step_size, Turn_off_80mA_HL_step_size, Rf_step_size, Rr_step_size
import os, shutil
import matplotlib.pyplot as plt
import matplotlib

####################################################################################################################################
#                                                         SERVER                                                                   #
####################################################################################################################################

# Set Flask server over here
app = Flask(__name__)
CORS(app)

# @app.before_request
# def initialize():
#     plt.figure(figsize=(10, 6))

# Retrieving current working directory
@app.route('/get-current-working-directory')
def getCurrWork():
    cwd = os.getcwd()
    return cwd

# Retrieving list of available folders in current working directory 
@app.route('/get-list-of-folders')
def getListOfDir():
    cwd = getCurrWork()
    listOfDir = os.scandir(cwd)
    listOfDir = check_files(listOfDir)
    return jsonify(listOfDir)

def check_files(list):
    result = []
    for file in list:
        if 'Processed' in file.name or file.name in processed or file.is_dir() == False or file.name == '.git' or file.name == '__pycache__' or file.name == '$RECYCLE.BIN' or file.name == '.ipynb_checkpoints' or file.name == '.jpeg File':
            pass
        else:
            result.append(file.name)
    return result

# Navigate backward one directory
@app.route('/directory-backward')
def directoryBackward():
    os.chdir('../')
    cwd = os.getcwd()
    return cwd

# Navigate forward one directory 
@app.route('/directory-forward', methods=['POST'])
def directoryForward():
    data = request.json
    selectedFolder = data.get('selectedFolder')
    path = f'./{selectedFolder}'
    os.chdir(path)
    cwd = os.getcwd()
    return cwd

# Set folder variable and create a respective folder in current working directory
@app.route('/set-folder-and-create', methods=['POST'])
def setFolderAndCreate():
    data = request.json
    selectedFolder = data.get('finalFolder')
    cwd = os.getcwd()
    processed.append(selectedFolder)
    selectedFolderPath = f'{cwd}/{selectedFolder}'
    fileDetail['selected_folder'] = selectedFolderPath
    finalFolderPath = f'{fileDetail["selected_folder"]}_Processed'
    fileDetail['fdpath'] = finalFolderPath
    os.mkdir(fileDetail['fdpath'])
    return finalFolderPath

# Set the selected columns from frontend user interface
@app.route('/set-selected-columns', methods=['POST'])
def setSelectedColumns():
    data = request.json
    selectedColumns = data.get('selectedColumns')
    # 1st loop reset the values to 0
    for key in fileDetail['selected_column']:
        fileDetail['selected_column'][key] = 0
    # 2nd loop set the values based on selectedColumns
    for row in list(selectedColumns.keys()):
        if (selectedColumns[row][1]):
            fileDetail['selected_column'][row] = 1
    setStepSize(selectedColumns)
    return selectedColumns

# Helper function for setSelectedColumns()
# Set step size for each selected column
def setStepSize(selectedColumns):
    if (selectedColumns['Isc_20mA'][1]):
        Isc_20mA_data['step_size_Isc_20mA'] = float(selectedColumns['Isc_20mA'][0])
    if (selectedColumns['Turn_off_80mA_'][1]):
        Turn_off_80mA_data['step_size_Turn_off_80mA'] = float(selectedColumns['Turn_off_80mA_'][0])
    if (selectedColumns['Turn_off_80mA_HL'][1]):
        Turn_off_80mA_HL_data['step_size_Turn_off_80mA_HL'] = float(selectedColumns['Turn_off_80mA_HL'][0])
    if (selectedColumns['Rf'][1]):
        Rf_data['step_size_Rf'] = float(selectedColumns['Rf'][0])
    if (selectedColumns['Rr'][1]):
        Rr_data['step_size_Rr'] = float(selectedColumns['Rr'][0])

# Set the upper and lower range limit for each column
@app.route('/set-selected-range', methods=['POST'])
def setSelectedRange():
    data = request.json
    selectedRange = data.get('selectedRange')
    if (fileDetail['selected_column']['Isc_20mA'] == 1):
        Isc_20mA_data['data_lLimit_Isc_20mA'] = float(selectedRange['Isc_20mA'][0])
        Isc_20mA_data['data_uLimit_Isc_20mA'] = float(selectedRange['Isc_20mA'][1])
    if (fileDetail['selected_column']['Turn_off_80mA_'] == 1):
        Turn_off_80mA_data['data_lLimit_Turn_off_80mA'] = float(selectedRange['Turn_off_80mA_'][0])
        Turn_off_80mA_data['data_uLimit_Turn_off_80mA'] = float(selectedRange['Turn_off_80mA_'][1])
    if (fileDetail['selected_column']['Turn_off_80mA_HL'] == 1):
        Turn_off_80mA_HL_data['data_lLimit_Turn_off_80mA_HL'] = float(selectedRange['Turn_off_80mA_HL'][0])
        Turn_off_80mA_HL_data['data_uLimit_Turn_off_80mA_HL'] = float(selectedRange['Turn_off_80mA_HL'][1])
    if (fileDetail['selected_column']['Rf'] == 1):
        Rf_data['data_lLimit_Rf'] = float(selectedRange['Rf'][0])
        Rf_data['data_uLimit_Rf'] = float(selectedRange['Rf'][1])
    if (fileDetail['selected_column']['Rr'] == 1):
        Rr_data['data_lLimit_Rr'] = float(selectedRange['Rr'][0])
        Rr_data['data_uLimit_Rr'] = float(selectedRange['Rr'][1])
    return selectedRange

# Determine whether the current process is fast-track / rush piece
@app.route('/set-fast-track', methods=['POST'])
def setFastTrack():
    data = request.json
    fastTrack = data.get('fastTrack')
    if fastTrack:
        fileDetail['fast_track'] = True
    else:
        fileDetail['fast_track'] = False
    return fastTrack

# This is the main execution function 
# It calls functions in helper, util, step, and main file
@app.route('/exec')
def exec():
    stage = []
    stage.append('**** Enter exec() in server ****')
    stage.append(fileDetail['selected_column'])
    try :
        stage.append(' - File Creation Start')
        if (fileDetail['selected_column']['Isc_20mA'] == 1):
            try:
                Isc_20_mA_step_size()
                # Create a new file and new folder for storing charts
                fileDetail['isc_fdpath'] = f'{fileDetail["fdpath"]}/Isc_20mA'
                os.mkdir(fileDetail['isc_fdpath'])
                fileDetail['isc_fdpath_wmap'] = f'{fileDetail["fdpath"]}/Isc_20mA_wmap'
                os.mkdir(fileDetail['isc_fdpath_wmap'])
                wb = Workbook()
                wb.save(f'{fileDetail["fdpath"]}/Isc_20mA.xlsx')
                wb_wmap = Workbook()
                wb_wmap.save(f'{fileDetail["fdpath"]}/Isc_20mA_wmap.xlsx')
                stage.append(' - Isc_20mA File Creation Complete')
            except:
                stage.append('Error in file creation: Error when creaing Isc_20mA files')

        if (fileDetail['selected_column']['Turn_off_80mA_'] == 1):
            try:
                Turn_off_80mA_step_size()
                # Create a new file and new folder for storing charts
                fileDetail['turn_off_fdpath'] = f'{fileDetail["fdpath"]}/Turn_off_80mA_'
                os.mkdir(fileDetail['turn_off_fdpath'])
                fileDetail['turn_off_fdpath_wmap'] = f'{fileDetail["fdpath"]}/Turn_off_80mA__wmap'
                os.mkdir(fileDetail['turn_off_fdpath_wmap'])
                wb = Workbook()
                wb.save(f'{fileDetail["fdpath"]}/Turn_off_80mA_.xlsx')
                wb_wmap = Workbook()
                wb_wmap.save(f'{fileDetail["fdpath"]}/Turn_off_80mA__wmap.xlsx')
                stage.append(' - Turn_off_80mA File Creation Complete')
            except:
                stage.append('Error in file creation: Error when creaing Turn_off_80mA_ files')

        if (fileDetail['selected_column']['Turn_off_80mA_HL'] == 1):
            try:
                Turn_off_80mA_HL_step_size()
                # Create a new file and new folder for storing charts
                fileDetail['turn_off_HL_fdpath'] = f'{fileDetail["fdpath"]}/Turn_off_80mA_HL'
                os.mkdir(fileDetail['turn_off_HL_fdpath'])
                fileDetail['turn_off_HL_fdpath_wmap'] = f'{fileDetail["fdpath"]}/Turn_off_80mA_HL_wmap'
                os.mkdir(fileDetail['turn_off_HL_fdpath_wmap'])
                wb = Workbook()
                wb.save(f'{fileDetail["fdpath"]}/Turn_off_80mA_HL.xlsx')
                wb_wmap = Workbook()
                wb_wmap.save(f'{fileDetail["fdpath"]}/Turn_off_80mA_HL_wmap.xlsx')
                stage.append(' - Turn_off_80mA_HL File Creation Complete')
            except:
                stage.append('Error in file creation: Error when creaing Turn_off_80mA_HL files')
        if (fileDetail['selected_column']['Rf'] == 1):
            try: 
                Rf_step_size()
                # Create a new file and new folder for storing charts
                fileDetail['rf_fdpath'] = f'{fileDetail["fdpath"]}/Rf'
                os.mkdir(fileDetail['rf_fdpath'])
                fileDetail['rf_fdpath_wmap'] = f'{fileDetail["fdpath"]}/Rf_wmap'
                os.mkdir(fileDetail['rf_fdpath_wmap'])
                wb = Workbook()
                wb.save(f'{fileDetail["fdpath"]}/Rf.xlsx')
                wb_wmap = Workbook()
                wb_wmap.save(f'{fileDetail["fdpath"]}/Rf_wmap.xlsx')
                stage.append(' - Rf File Creation Complete')
            except:
                stage.append('Error in file creation: Error when creaing Rf files')
        if (fileDetail['selected_column']['Rr'] == 1):
            try:
                Rr_step_size()  
                # Create a new file and new folder for storing charts
                fileDetail['rr_fdpath'] = f'{fileDetail["fdpath"]}/Rr'
                os.mkdir(fileDetail['rr_fdpath'])
                fileDetail['rr_fdpath_wmap'] = f'{fileDetail["fdpath"]}/Rr_wmap'
                os.mkdir(fileDetail['rr_fdpath_wmap'])
                wb = Workbook()
                wb.save(f'{fileDetail["fdpath"]}/Rr.xlsx')
                wb_wmap = Workbook()
                wb_wmap.save(f'{fileDetail["fdpath"]}/Rr_wmap.xlsx')
                stage.append(' - Rr File Creation Complete')
            except:
                stage.append('Error in file creation: Error when creaing Rr files')
    except:
        stage.append('Error in file creation: Error locate in first half of exec() in server.py')
    
    matplotlib.use('agg')
    try:
        stage.append(' - Calling reading_and_processing()')
        stage += reading_and_processing()
    except:
        stage.append('Error when preparing to enter reading_and_processing(): Error occur because of calling reading_and_processing() or in the function itself in server.py or main.py')
    stage.append('SUCCESSFUL')
    return stage

# Retrieve the current progress of the process, sends the progress back to frontend
@app.route('/get-progress')
def getProgress():
    count = progressTracker['count']
    progressStatements = progressTracker['progressStatements']
    result = [count, progressStatements]
    return result

# Reset the variable processed, so all folders are visible in the folder selection tab
@app.route('/reset-variable-processed-and-remove-all-processed-folder')
def resetVariableProcessedAndDelete():
    processed.clear()
    cwd = getCurrWork() 
    listOfDir = os.scandir(cwd)
    for folder in listOfDir:
        if "Processed" in folder.name:
            folderPath = os.path.join(cwd, folder.name)
            shutil.rmtree(folderPath)
    return 'Variable Processed Reset'

# Remove the created content in the folder but doesn't delete the folder if the process is canceled half way
@app.route('/remove-processed-folder-content')
def removeProcessedFolderContent():
    finalFolderPath = fileDetail['fdpath']
    shutil.rmtree(finalFolderPath)
    os.mkdir(finalFolderPath)
    return 'Folder Content Deleted'

# Remove the folder entirely 
@app.route('/remove-folder')
def removeFolder():
    finalFolderPath = fileDetail['fdpath']
    processed.pop()
    os.rmdir(finalFolderPath)
    return 'Folder Deleted'

# Retrieve new folder path
@app.route('/get-new-folder-path')
def getNewFolderPath():
    finalFolderPath = fileDetail['fdpath']
    return finalFolderPath

# Reset all used variables
@app.route('/reset-all-data')
def resetAllData():
    fileDetail = {
        'selected_folder': '',
        'all_column_headers': [
            'Isc_20mA', 
            'Turn_off_80mA_', 
            'Turn_off_80mA_HL', 
            'Rf', 
            'Rr'
        ],
        'selected_column': {
            'Isc_20mA': 0,
            'Turn_off_80mA_': 0,
            'Turn_off_80mA_HL': 0,
            'Rf': 0,
            "Rr": 0,
        },
        'fdpath': '',
        'isc_fdpath': '',
        'isc_fdpath_wmap': '',
        'turn_off_fdpath': '',
        'turn_off_fdpath_wmap': '',
        'turn_off_HL_fdpath': '',
        'turn_off_HL_fdpath_wmap': '',
        'rf_fdpath': '',
        'rf_fdpath_wmap': '',
        'rr_fdpath': '',
        'rr_fdpath_wmap': '',
        'all_files': [],
        'file_download_path': '',
        'first_df': {},
        'second_df': {},
        'df': {},
        'fast_track': '',
    }
    progressTracker = {
        'count': 0,
        'progressStatements': []
    }
    processed = []
    Isc_20mA_data = {
        'step_size_Isc_20mA': 0.15,
        'len_Isc_20mA': 0,
        'x_Isc_20mA': [],
        'y_Isc_20mA': [],
        'uLimit_Isc_20mA': 10,
        'lLimit_Isc_20mA': 2,
        'data_uLimit_Isc_20mA': 10,
        'data_lLimit_Isc_20mA': 0,
    }
    Turn_off_80mA_data = {
        'step_size_Turn_off_80mA': 0.8,
        'len_Turn_off_80mA': 0,
        'x_Turn_off_80mA': [],
        'y_Turn_off_80mA': [],
        'uLimit_Turn_off_80mA': 500,
        'lLimit_Turn_off_80mA': 0,
        'data_uLimit_Turn_off_80mA': 60,
        'data_lLimit_Turn_off_80mA': 0,
    }

    Turn_off_80mA_HL_data = {
        'step_size_Turn_off_80mA_HL': 0.5,
        'len_Turn_off_80mA_HL': 0,
        'x_Turn_off_80mA_HL': [],
        'y_Turn_off_80mA_HL': [],
        'uLimit_Turn_off_80mA_HL': 20,
        'lLimit_Turn_off_80mA_HL': -20,
        'data_uLimit_Turn_off_80mA_HL': 10,
        'data_lLimit_Turn_off_80mA_HL': -10,
    }
    Rf_data = {
        'step_size_Rf': 0.2,
        'len_Rf': 0,
        'x_Rf': [],
        'y_Rf': [],
        'uLimit_Rf': 20,
        'lLimit_Rf': 10,
        'data_uLimit_Rf': 20,
        'data_lLimit_Rf': 10,
    }
    Rr_data = {
        'step_size_Rr': 0.2,
        'len_Rr': 0,
        'x_Rr': [],
        'y_Rr': [],
        'uLimit_Rr': 20,
        'lLimit_Rr': 10,
        'data_uLimit_Rr': 20,
        'data_lLimit_Rr': 10,
    }
    result = 'All Variables reset'
    return result

############################################### Routes below are for testing purposes ###############################################
@app.route('/set-test', methods=['POST'])
def testRoute():
    data = request.json.get('data')
    return data

@app.route('/get-test')
def getTest():
    result = 'Test Successful'
    return result

@app.route('/check-file-detail')
def checkFileDetail():
    result = []
    result.append(fileDetail['selected_folder'])
    result.append(fileDetail['fdpath'])
    result.append(fileDetail['selected_column'])
    result.append(fileDetail['fast_track'])
    return result

@app.route('/test-plt')
def testPlt():
    matplotlib.use('agg')
    path = '/Users/wng/Desktop'
    x_values = [1, 2, 3, 4, 5]
    y_values = [2, 4, 6, 8, 10]

    # Create the plot
    plt.plot(x_values, y_values)

    # Add title and labels
    plt.title('Simple Line Chart')
    plt.xlabel('X-axis Label')
    plt.ylabel('Y-axis Label')

    # Display the plot
    plt.savefig(path + '/test.png')

    return 'sucessful'

if __name__ == '__main__':
    app.run(debug=True)
    print('All server code executed_1')

print('All server code executed_2')